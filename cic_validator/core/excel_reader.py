"""Read Excel masterfile rows into Record objects."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from cic_validator.core.models import Record
from cic_validator.core.schema_loader import SchemaLoader


def read_excel_file(
    path: str | Path,
    schema: SchemaLoader,
) -> list[Record]:
    p = Path(path)
    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel workbook has no active sheet")

    records: list[Record] = []
    line_number = 0
    for row in ws.iter_rows(values_only=True):
        line_number += 1
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        rt = str(row[0]).strip() if row[0] is not None else ""
        if rt not in schema.record_types():
            # Treat as a header/garbage row if the first cell is not a record type.
            continue
        expected = schema.get_field_count(rt)
        values = [str(v).strip() if v is not None else "" for v in row[1:]]
        # Pad or trim to expected count.
        if len(values) < expected - 1:
            values.extend([""] * (expected - 1 - len(values)))
        values = values[: expected - 1]
        parts = [rt] + values

        field_names = [f["name"] for f in schema.get_fields(rt)]
        fields = {}
        for i, name in enumerate(field_names):
            fields[name] = parts[i] if i < len(parts) else ""
        psn = fields.get("Provider Subject No") or None
        records.append(
            Record(record_type=rt, line_number=line_number, provider_subject_no=psn, fields=fields, raw_field_count=len(parts))
        )

    return records
